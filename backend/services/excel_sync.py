"""Excel synchronization service."""

import shutil
import re
import requests
from datetime import datetime
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse, parse_qs

from openpyxl import load_workbook

from config import EXCEL_FILE_PATH, BACKUP_DIR, SHAREPOINT_EXCEL_URL
from database import get_db
from models import SyncResult


class ExcelSyncService:
    """Service for syncing data between Excel and SQLite database."""

    def __init__(self):
        self.excel_path = EXCEL_FILE_PATH
        self.backup_dir = BACKUP_DIR
        self.sharepoint_url = SHAREPOINT_EXCEL_URL

    def _convert_sharepoint_to_download_url(self, sharing_url: str) -> str:
        """Convert a OneDrive/SharePoint sharing URL to a direct download URL."""
        # Parse the sharing URL
        # Format: https://domain-my.sharepoint.com/:x:/g/personal/user/ID?e=token
        # Download: https://domain-my.sharepoint.com/personal/user/_layouts/15/download.aspx?share=ID

        parsed = urlparse(sharing_url)

        # Extract the path components
        # /:x:/g/personal/zach_b_ellingsonmotorcars_com/IQCA6...
        path_match = re.match(r'/:x:/g/personal/([^/]+)/([^?]+)', parsed.path)
        if path_match:
            user = path_match.group(1)
            file_id = path_match.group(2)
            download_url = f"{parsed.scheme}://{parsed.netloc}/personal/{user}/_layouts/15/download.aspx?share={file_id}"
            return download_url

        # If pattern doesn't match, try alternate approach - add download=1 parameter
        if '?' in sharing_url:
            return sharing_url + '&download=1'
        return sharing_url + '?download=1'

    def _download_from_sharepoint(self) -> Optional[Path]:
        """Download Excel file from SharePoint/OneDrive and return local path."""
        if not self.sharepoint_url:
            return None

        try:
            download_url = self._convert_sharepoint_to_download_url(self.sharepoint_url)

            # Download the file
            response = requests.get(download_url, allow_redirects=True, timeout=60)
            response.raise_for_status()

            # Save to temp file
            temp_path = self.backup_dir / "sharepoint_download.xlsx"
            with open(temp_path, 'wb') as f:
                f.write(response.content)

            return temp_path

        except Exception as e:
            print(f"Failed to download from SharePoint: {e}")
            return None

    def _create_backup(self) -> Path:
        """Create a backup of the Excel file."""
        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"LoadBoard_backup_{timestamp}.xlsx"
        backup_path = self.backup_dir / backup_name
        shutil.copy2(self.excel_path, backup_path)
        return backup_path

    def _parse_date(self, value) -> Optional[str]:
        """Parse date value from Excel to ISO format string."""
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            # Try common date formats
            for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"]:
                try:
                    return datetime.strptime(value, fmt).date().isoformat()
                except ValueError:
                    continue
        return None

    def _parse_bool(self, value) -> bool:
        """Parse boolean value from Excel."""
        if value is None:
            return False
        if isinstance(value, bool):
            return value
        if isinstance(value, str):
            return value.strip().upper() in ("YES", "Y", "TRUE", "1", "X", "YES,DELAYED")
        if isinstance(value, (int, float)):
            return bool(value)
        return False

    def _parse_shipped_status(self, value) -> tuple[bool, bool]:
        """Parse shipped status from Excel. Returns (shipped, delayed) tuple.

        - 'No' or empty -> (False, False) - not shipped
        - 'Yes' -> (True, False) - shipped on time
        - 'Yes-Delayed' or 'Yes,Delayed' -> (True, True) - shipped but late
        """
        if value is None:
            return (False, False)
        if isinstance(value, bool):
            return (value, False)
        if isinstance(value, str):
            val = value.strip().upper()
            if val in ("YES-DELAYED", "YES,DELAYED", "YES, DELAYED", "YES-LATE", "YES,LATE", "YES, LATE"):
                return (True, True)  # Shipped but delayed
            elif val in ("YES", "Y", "TRUE", "1", "X"):
                return (True, False)  # Shipped on time
        if isinstance(value, (int, float)):
            return (bool(value), False)
        return (False, False)

    def _parse_number(self, value) -> Optional[float]:
        """Parse numeric value from Excel."""
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None
            try:
                return float(value)
            except ValueError:
                return None
        return None

    def _log_sync(self, sync_type: str, status: str, records: int, details: str = None):
        """Log sync operation to database."""
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO sync_log (sync_type, status, records_processed, details)
                VALUES (?, ?, ?, ?)
            """, (sync_type, status, records, details))
            conn.commit()

    def import_from_excel(self) -> SyncResult:
        """Import data from Excel file into the database."""
        errors = []
        records_processed = 0
        wb = None
        temp_path = None
        downloaded_from_sharepoint = False

        try:
            # Try SharePoint first
            if self.sharepoint_url:
                temp_path = self._download_from_sharepoint()
                if temp_path and temp_path.exists():
                    downloaded_from_sharepoint = True

            # Fall back to local file
            if not temp_path or not temp_path.exists():
                if not self.excel_path.exists():
                    return SyncResult(
                        success=False,
                        message=f"Excel file not found: {self.excel_path}",
                        errors=[f"File not found: {self.excel_path}"]
                    )
                temp_path = self.backup_dir / "temp_import.xlsx"
                # Create a temporary copy to avoid lock issues
                shutil.copy2(self.excel_path, temp_path)

            wb = load_workbook(temp_path, read_only=True, data_only=True)

            with get_db() as conn:
                cursor = conn.cursor()

                # Import TP INBOUND
                if "TP INBOUND" in wb.sheetnames:
                    sheet = wb["TP INBOUND"]
                    count = self._import_inbound_sheet(cursor, sheet, "TP")
                    records_processed += count

                # Import OTHERINBOUND
                if "OTHERINBOUND" in wb.sheetnames:
                    sheet = wb["OTHERINBOUND"]
                    count = self._import_inbound_sheet(cursor, sheet, "OTHER")
                    records_processed += count

                # Import TP OUTBOUND
                if "TP OUTBOUND" in wb.sheetnames:
                    sheet = wb["TP OUTBOUND"]
                    count = self._import_outbound_sheet(cursor, sheet, "TP")
                    records_processed += count

                # Import OTHEROUTBOUND
                if "OTHEROUTBOUND" in wb.sheetnames:
                    sheet = wb["OTHEROUTBOUND"]
                    count = self._import_outbound_sheet(cursor, sheet, "OTHER")
                    records_processed += count

                # Import Carriers & Customers
                if "Carriers&Customers" in wb.sheetnames:
                    sheet = wb["Carriers&Customers"]
                    count = self._import_reference_sheet(cursor, sheet)
                    records_processed += count

                # Import Products
                if "Product Counts" in wb.sheetnames:
                    sheet = wb["Product Counts"]
                    count = self._import_products_sheet(cursor, sheet)
                    records_processed += count

                conn.commit()

            source_msg = "from SharePoint" if downloaded_from_sharepoint else "from local file"
            self._log_sync("import", "success", records_processed, source_msg)

            return SyncResult(
                success=True,
                message=f"Successfully imported {records_processed} records {source_msg}",
                records_processed=records_processed,
                errors=errors
            )

        except PermissionError as e:
            errors.append(f"File is locked: {str(e)}")
            self._log_sync("import", "error", 0, f"Permission denied: {str(e)}")
            return SyncResult(
                success=False,
                message="Import failed: The Excel file is open in another application. Please close Excel and try again.",
                errors=errors
            )
        except Exception as e:
            errors.append(str(e))
            self._log_sync("import", "error", 0, str(e))
            return SyncResult(
                success=False,
                message=f"Import failed: {str(e)}",
                errors=errors
            )
        finally:
            # Always close workbook if it was opened
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
            # Clean up temp file
            temp_path.unlink(missing_ok=True)

    def _import_inbound_sheet(self, cursor, sheet, source: str) -> int:
        """Import inbound shipments from a sheet."""
        count = 0
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        for row_num, row in enumerate(rows, start=2):
            if not row or all(cell is None for cell in row):
                continue

            # Column mapping based on Excel structure
            # TP INBOUND: Item #, Cases, PO, Carrier, BOL #, TP Receipt #, Date, Received, Pallets, Notes
            # OTHERINBOUND: Item #, Cases, PO, Carrier, BOL #, Date, Received, Pallets, Notes
            if source == "TP":
                item_number = str(row[0]) if row[0] else None
                cases = self._parse_number(row[1])
                po = str(row[2]) if row[2] else None
                carrier = str(row[3]) if row[3] else None
                bol_number = str(row[4]) if row[4] else None
                tp_receipt = str(row[5]) if row[5] else None
                ship_date = self._parse_date(row[6])
                received = self._parse_bool(row[7])
                pallets = self._parse_number(row[8])
                notes = str(row[9]) if len(row) > 9 and row[9] else None
            else:
                item_number = str(row[0]) if row[0] else None
                cases = self._parse_number(row[1])
                po = str(row[2]) if row[2] else None
                carrier = str(row[3]) if row[3] else None
                bol_number = str(row[4]) if row[4] else None
                tp_receipt = None
                ship_date = self._parse_date(row[5])
                received = self._parse_bool(row[6])
                pallets = self._parse_number(row[7])
                notes = str(row[8]) if len(row) > 8 and row[8] else None

            # Check if record exists by excel_row
            cursor.execute(
                "SELECT id FROM inbound_shipments WHERE source = ? AND excel_row = ?",
                (source, row_num)
            )
            existing = cursor.fetchone()

            now = datetime.now().isoformat()

            if existing:
                cursor.execute("""
                    UPDATE inbound_shipments SET
                        item_number = ?, cases = ?, po = ?, carrier = ?,
                        bol_number = ?, tp_receipt_number = ?, ship_date = ?,
                        received = ?, pallets = ?, notes = ?,
                        updated_at = ?, synced_at = ?
                    WHERE id = ?
                """, (
                    item_number, int(cases) if cases else None, po, carrier,
                    bol_number, tp_receipt, ship_date,
                    1 if received else 0, pallets, notes,
                    now, now, existing[0]
                ))
            else:
                cursor.execute("""
                    INSERT INTO inbound_shipments (
                        source, item_number, cases, po, carrier, bol_number,
                        tp_receipt_number, ship_date, received, pallets, notes,
                        excel_row, created_at, updated_at, synced_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    source, item_number, int(cases) if cases else None, po, carrier,
                    bol_number, tp_receipt, ship_date, 1 if received else 0,
                    pallets, notes, row_num, now, now, now
                ))

            count += 1

        return count

    def _import_outbound_sheet(self, cursor, sheet, source: str) -> int:
        """Import outbound shipments from a sheet."""
        count = 0
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        for row_num, row in enumerate(rows, start=2):
            if not row or all(cell is None for cell in row):
                continue

            # Column mapping:
            # TP OUTBOUND: Reference #, Order #, Customer, Ship Date, Carrier, Shipped, Pallets, Pro, Seal, Notes, Time
            # OTHEROUTBOUND: Reference #, Order #, Customer, Ship Date, Carrier, Shipped, Actual Date, Pallets, Pro, Seal, Notes
            reference_number = str(row[0]) if row[0] else None
            order_number = str(row[1]) if row[1] else None

            # Skip rows without order_number (column F in Excel)
            if not order_number:
                continue
            customer = str(row[2]) if row[2] else None
            ship_date = self._parse_date(row[3])
            carrier = str(row[4]) if row[4] else None
            shipped, delayed = self._parse_shipped_status(row[5])

            if source == "TP":
                actual_date = None
                pallets = self._parse_number(row[6])
                pro = str(row[7]) if len(row) > 7 and row[7] else None
                seal = str(row[8]) if len(row) > 8 and row[8] else None
                notes = str(row[9]) if len(row) > 9 and row[9] else None
                pickup_time = str(row[10]) if len(row) > 10 and row[10] else None
            else:
                actual_date = self._parse_date(row[6])
                pallets = self._parse_number(row[7])
                pro = str(row[8]) if len(row) > 8 and row[8] else None
                seal = str(row[9]) if len(row) > 9 and row[9] else None
                notes = str(row[10]) if len(row) > 10 and row[10] else None
                pickup_time = None

            # Check if record exists
            cursor.execute(
                "SELECT id FROM outbound_shipments WHERE source = ? AND excel_row = ?",
                (source, row_num)
            )
            existing = cursor.fetchone()

            now = datetime.now().isoformat()

            if existing:
                cursor.execute("""
                    UPDATE outbound_shipments SET
                        reference_number = ?, order_number = ?, customer = ?,
                        ship_date = ?, carrier = ?, shipped = ?, delayed = ?, actual_date = ?,
                        pallets = ?, pro = ?, seal = ?, notes = ?, pickup_time = ?,
                        updated_at = ?, synced_at = ?
                    WHERE id = ?
                """, (
                    reference_number, order_number, customer, ship_date,
                    carrier, 1 if shipped else 0, 1 if delayed else 0, actual_date, pallets,
                    pro, seal, notes, pickup_time, now, now, existing[0]
                ))
            else:
                cursor.execute("""
                    INSERT INTO outbound_shipments (
                        source, reference_number, order_number, customer,
                        ship_date, carrier, shipped, delayed, actual_date, pallets,
                        pro, seal, notes, pickup_time, excel_row,
                        created_at, updated_at, synced_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    source, reference_number, order_number, customer,
                    ship_date, carrier, 1 if shipped else 0, 1 if delayed else 0, actual_date,
                    pallets, pro, seal, notes, pickup_time, row_num,
                    now, now, now
                ))

            count += 1

        return count

    def _import_reference_sheet(self, cursor, sheet) -> int:
        """Import carriers and customers from reference sheet."""
        count = 0
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not row:
                continue

            # Column A: Carriers, Column B: Customers (typical layout)
            carrier = str(row[0]).strip() if row[0] else None
            customer = str(row[1]).strip() if len(row) > 1 and row[1] else None

            if carrier:
                cursor.execute(
                    "INSERT OR IGNORE INTO carriers (name) VALUES (?)",
                    (carrier,)
                )
                count += 1

            if customer:
                cursor.execute(
                    "INSERT OR IGNORE INTO customers (name) VALUES (?)",
                    (customer,)
                )
                count += 1

        return count

    def _import_products_sheet(self, cursor, sheet) -> int:
        """Import products from product counts sheet."""
        count = 0
        rows = list(sheet.iter_rows(min_row=2, values_only=True))

        for row in rows:
            if not row or not row[0]:
                continue

            item_number = str(row[0]).strip()

            # Check if product exists
            cursor.execute(
                "SELECT id FROM products WHERE item_number = ?",
                (item_number,)
            )
            existing = cursor.fetchone()

            # Parse product data
            items_per_case = int(self._parse_number(row[1]) or 0) if len(row) > 1 else None
            items_per_pallet = int(self._parse_number(row[2]) or 0) if len(row) > 2 else None
            cases_per_pallet = int(self._parse_number(row[3]) or 0) if len(row) > 3 else None
            layers_per_pallet = int(self._parse_number(row[4]) or 0) if len(row) > 4 else None
            cases_per_layer = int(self._parse_number(row[5]) or 0) if len(row) > 5 else None
            notes = str(row[6]) if len(row) > 6 and row[6] else None

            if existing:
                cursor.execute("""
                    UPDATE products SET
                        items_per_case = ?, items_per_pallet = ?,
                        cases_per_pallet = ?, layers_per_pallet = ?,
                        cases_per_layer = ?, notes = ?
                    WHERE id = ?
                """, (
                    items_per_case, items_per_pallet, cases_per_pallet,
                    layers_per_pallet, cases_per_layer, notes, existing[0]
                ))
            else:
                cursor.execute("""
                    INSERT INTO products (
                        item_number, items_per_case, items_per_pallet,
                        cases_per_pallet, layers_per_pallet, cases_per_layer, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    item_number, items_per_case, items_per_pallet,
                    cases_per_pallet, layers_per_pallet, cases_per_layer, notes
                ))

            count += 1

        return count

    def _wait_for_file_access(self, file_path: Path, max_retries: int = 5, delay: float = 1.0) -> bool:
        """Wait for file to become accessible (not locked by another process)."""
        import time
        for attempt in range(max_retries):
            try:
                # Try to open file in append mode to check if it's locked
                with open(file_path, 'a'):
                    pass
                return True
            except (IOError, PermissionError):
                if attempt < max_retries - 1:
                    time.sleep(delay)
        return False

    def export_to_excel(self) -> SyncResult:
        """Export database changes to the Excel file."""
        if not self.excel_path.exists():
            return SyncResult(
                success=False,
                message=f"Excel file not found: {self.excel_path}",
                errors=[f"File not found: {self.excel_path}"]
            )

        errors = []
        records_processed = 0
        wb = None
        temp_path = self.backup_dir / "temp_export.xlsx"

        try:
            # Create backup first
            self._create_backup()

            # Copy to temp file to avoid locking the original
            shutil.copy2(self.excel_path, temp_path)

            # Work with the temp copy
            wb = load_workbook(temp_path)

            with get_db() as conn:
                cursor = conn.cursor()

                # Export TP INBOUND
                if "TP INBOUND" in wb.sheetnames:
                    sheet = wb["TP INBOUND"]
                    count = self._export_inbound_sheet(cursor, sheet, "TP")
                    records_processed += count

                # Export OTHERINBOUND
                if "OTHERINBOUND" in wb.sheetnames:
                    sheet = wb["OTHERINBOUND"]
                    count = self._export_inbound_sheet(cursor, sheet, "OTHER")
                    records_processed += count

                # Export TP OUTBOUND
                if "TP OUTBOUND" in wb.sheetnames:
                    sheet = wb["TP OUTBOUND"]
                    count = self._export_outbound_sheet(cursor, sheet, "TP")
                    records_processed += count

                # Export OTHEROUTBOUND
                if "OTHEROUTBOUND" in wb.sheetnames:
                    sheet = wb["OTHEROUTBOUND"]
                    count = self._export_outbound_sheet(cursor, sheet, "OTHER")
                    records_processed += count

            # Save to temp file first
            wb.save(temp_path)
            wb.close()
            wb = None

            # Check if original file is accessible before replacing
            if not self._wait_for_file_access(self.excel_path):
                # Can't access original, save to alternate location
                alt_path = self.backup_dir / f"LoadBoard_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                shutil.copy2(temp_path, alt_path)
                temp_path.unlink(missing_ok=True)
                return SyncResult(
                    success=True,
                    message=f"Original file locked. Exported {records_processed} records to: {alt_path}",
                    records_processed=records_processed,
                    errors=["Original file was locked, saved to backup location"]
                )

            # Replace original with updated temp file
            shutil.copy2(temp_path, self.excel_path)
            temp_path.unlink(missing_ok=True)

            # Update synced_at for all records
            now = datetime.now().isoformat()
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute("UPDATE inbound_shipments SET synced_at = ?", (now,))
                cursor.execute("UPDATE outbound_shipments SET synced_at = ?", (now,))
                conn.commit()

            self._log_sync("export", "success", records_processed)

            return SyncResult(
                success=True,
                message=f"Successfully exported {records_processed} records",
                records_processed=records_processed,
                errors=errors
            )

        except PermissionError as e:
            errors.append(f"File is locked: {str(e)}")
            self._log_sync("export", "error", 0, f"Permission denied: {str(e)}")
            return SyncResult(
                success=False,
                message="Export failed: The Excel file is open in another application. Please close Excel and try again.",
                errors=errors
            )
        except Exception as e:
            errors.append(str(e))
            self._log_sync("export", "error", 0, str(e))
            return SyncResult(
                success=False,
                message=f"Export failed: {str(e)}",
                errors=errors
            )
        finally:
            # Always close workbook if it was opened
            if wb is not None:
                try:
                    wb.close()
                except:
                    pass
            # Clean up temp file
            temp_path.unlink(missing_ok=True)

    def _export_inbound_sheet(self, cursor, sheet, source: str) -> int:
        """Export inbound shipments to Excel sheet."""
        count = 0

        # Get records that have been modified since last sync
        cursor.execute("""
            SELECT * FROM inbound_shipments
            WHERE source = ? AND excel_row IS NOT NULL
        """, (source,))
        rows = cursor.fetchall()

        for row in rows:
            excel_row = row["excel_row"]

            # Update cells in Excel based on source
            if source == "TP":
                sheet.cell(row=excel_row, column=1, value=row["item_number"])
                sheet.cell(row=excel_row, column=2, value=row["cases"])
                sheet.cell(row=excel_row, column=3, value=row["po"])
                sheet.cell(row=excel_row, column=4, value=row["carrier"])
                sheet.cell(row=excel_row, column=5, value=row["bol_number"])
                sheet.cell(row=excel_row, column=6, value=row["tp_receipt_number"])
                sheet.cell(row=excel_row, column=7, value=row["ship_date"])
                sheet.cell(row=excel_row, column=8, value="Yes" if row["received"] else "")
                sheet.cell(row=excel_row, column=9, value=row["pallets"])
                sheet.cell(row=excel_row, column=10, value=row["notes"])
            else:
                sheet.cell(row=excel_row, column=1, value=row["item_number"])
                sheet.cell(row=excel_row, column=2, value=row["cases"])
                sheet.cell(row=excel_row, column=3, value=row["po"])
                sheet.cell(row=excel_row, column=4, value=row["carrier"])
                sheet.cell(row=excel_row, column=5, value=row["bol_number"])
                sheet.cell(row=excel_row, column=6, value=row["ship_date"])
                sheet.cell(row=excel_row, column=7, value="Yes" if row["received"] else "")
                sheet.cell(row=excel_row, column=8, value=row["pallets"])
                sheet.cell(row=excel_row, column=9, value=row["notes"])

            count += 1

        return count

    def _export_outbound_sheet(self, cursor, sheet, source: str) -> int:
        """Export outbound shipments to Excel sheet."""
        count = 0

        cursor.execute("""
            SELECT * FROM outbound_shipments
            WHERE source = ? AND excel_row IS NOT NULL
        """, (source,))
        rows = cursor.fetchall()

        for row in rows:
            excel_row = row["excel_row"]

            sheet.cell(row=excel_row, column=1, value=row["reference_number"])
            sheet.cell(row=excel_row, column=2, value=row["order_number"])
            sheet.cell(row=excel_row, column=3, value=row["customer"])
            sheet.cell(row=excel_row, column=4, value=row["ship_date"])
            sheet.cell(row=excel_row, column=5, value=row["carrier"])
            sheet.cell(row=excel_row, column=6, value="Yes" if row["shipped"] else "")

            if source == "TP":
                sheet.cell(row=excel_row, column=7, value=row["pallets"])
                sheet.cell(row=excel_row, column=8, value=row["pro"])
                sheet.cell(row=excel_row, column=9, value=row["seal"])
                sheet.cell(row=excel_row, column=10, value=row["notes"])
                sheet.cell(row=excel_row, column=11, value=row["pickup_time"])
            else:
                sheet.cell(row=excel_row, column=7, value=row["actual_date"])
                sheet.cell(row=excel_row, column=8, value=row["pallets"])
                sheet.cell(row=excel_row, column=9, value=row["pro"])
                sheet.cell(row=excel_row, column=10, value=row["seal"])
                sheet.cell(row=excel_row, column=11, value=row["notes"])

            count += 1

        return count
